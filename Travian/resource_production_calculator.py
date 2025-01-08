import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "資源產能試算"

# 定義列寬
column_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 15, 'G': 20, 'H': 25}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 添加標題
ws.append([
    "編號", "資源類型", "當前等級", "基礎產能", "加乘後產能", "該級成本", "前一級產能", "回本時間(小時)"
])

# 標題置中並加粗
for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(bold=True)  # 設置粗體

# 定義資源類型和等級範圍
resource_types = ['木', '磚', '鐵', '米']
resource_list = ",".join(resource_types)
levels = list(range(0, 21))
level_list = ",".join(map(str, levels))

# 資源產能數據
resource_production_data = {
    0: [1, 1, 1, 1],
    1: [7, 7, 7, 7],
    2: [13, 13, 13, 13],
    3: [21, 21, 21, 21],
    4: [31, 31, 31, 31],
    5: [46, 46, 46, 46],
    6: [70, 70, 70, 70],
    7: [98, 98, 98, 98],
    8: [140, 140, 140, 140],
    9: [203, 203, 203, 203],
    10: [280, 280, 280, 280],
    11: [392, 392, 392, 392],
    12: [525, 525, 525, 525],
    13: [693, 693, 693, 693],
    14: [889, 889, 889, 889],
    15: [1120, 1120, 1120, 1120],
    16: [1400, 1400, 1400, 1400],
    17: [1820, 1820, 1820, 1820],
    18: [2240, 2240, 2240, 2240],
    19: [2800, 2800, 2800, 2800],
    20: [3430, 3430, 3430, 3430]
}

# 資源升級成本
resource_upgrade_costs = {
    0: [0, 0, 0, 0],
    1: [250, 250, 250, 250],   
    2: [415, 420, 450, 415],
    3: [695, 700, 755, 695],
    4: [1165, 1170, 1260, 1165],
    5: [1945, 1940, 2100, 1945],
    6: [3250, 3250, 3510, 3250],
    7: [5425, 5425, 5855, 5425],
    8: [9060, 9060, 9785, 9055],
    9: [12125, 12125, 16335, 12125],
    10: [25255, 25250, 27275, 25255],
    11: [42180, 42185, 45555, 42180],
    12: [70440, 70440, 76075, 70445],
    13: [117630, 117635, 127045, 117640],
    14: [196450, 196450, 212170, 196445],
    15: [328075, 328075, 354325, 328070],
    16: [547880, 547875, 591710, 547880],
    17: [914965, 914970, 988160, 914960],
    18: [1527990, 1527990, 1650225, 1527985],
    19: [2551745, 2551740, 2755880, 2551735],
    20: [4261410, 4261405, 4602325, 4261410]
}

# 創建資源產能和升級成本表，並隱藏
production_sheet = wb.create_sheet(title="ResourceProductionData")
upgrade_sheet = wb.create_sheet(title="BuildingUpgradeCosts")
for sheet, data, title in [(production_sheet, resource_production_data, "資源產能"), (upgrade_sheet, resource_upgrade_costs, "升級成本")]:
    sheet.append(["等級", "木", "磚", "鐵", "米"])
    for level, values in data.items():
        sheet.append([level] + values)
    sheet.sheet_state = 'hidden'

# 逐行插入數據並設置公式
for i in range(1, 19):
    ws.cell(row=i+1, column=1, value=i)

    dv_resource = DataValidation(type="list", formula1=f'"{resource_list}"', showDropDown=True)
    ws.add_data_validation(dv_resource)
    cell_resource = ws.cell(row=i+1, column=2, value="米")
    dv_resource.add(cell_resource)

    dv_level = DataValidation(type="list", formula1=f'"{level_list}"', showDropDown=True)
    ws.add_data_validation(dv_level)
    cell_level = ws.cell(row=i+1, column=3, value=20)
    dv_level.add(cell_level)

    # 基礎產能
    ws.cell(row=i+1, column=4, value=f'=INDEX(ResourceProductionData!E:E, C{i+1}+2)')

    # 加乘後產能
    ws.cell(row=i+1, column=5, value=(
        f'=(D{i+1} + '
        f'IF(B{i+1}="木", IF($B$23="是", 0.25*D{i+1}, 0) + IF($B$24>0, $B$24*D{i+1}, 0), '
        f'IF(B{i+1}="磚", IF($B$26="是", 0.25*D{i+1}, 0) + IF($B$27>0, $B$27*D{i+1}, 0), '
        f'IF(B{i+1}="鐵", IF($B$29="是", 0.25*D{i+1}, 0) + IF($B$30>0, $B$30*D{i+1}, 0), '
        f'IF(B{i+1}="米", IF($B$32="是", 0.25*D{i+1}, 0) + IF($B$33="是", 0.25*D{i+1}, 0) + IF($B$34>0, $B$34*D{i+1}, 0), 0))))'
        f') * IF(B{i+1}="木", IF($B$25="是", 1.25, 1), '
        f'IF(B{i+1}="磚", IF($B$28="是", 1.25, 1), '
        f'IF(B{i+1}="鐵", IF($B$31="是", 1.25, 1), '
        f'IF(B{i+1}="米", IF($B$35="是", 1.25, 1), 1))))'
    ))

    # 該級成本
    ws.cell(row=i+1, column=6, value=(
        f'=IF(B{i+1}="木", INDEX(BuildingUpgradeCosts!B$2:B$22, C{i+1}+1), '
        f'IF(B{i+1}="磚", INDEX(BuildingUpgradeCosts!C$2:C$22, C{i+1}+1), '
        f'IF(B{i+1}="鐵", INDEX(BuildingUpgradeCosts!D$2:D$22, C{i+1}+1), '
        f'IF(B{i+1}="米", INDEX(BuildingUpgradeCosts!E$2:E$22, C{i+1}+1)))))'
    ))

    # 前一級加乘後產能
    previous_level_production = ws.cell(row=i+1, column=7, value=(
        f'=IF(C{i+1}=0, 0, '
        f'IF(ISBLANK(INDEX(ResourceProductionData!E:E, C{i+1}+1)), 0, '
        f'INDEX(ResourceProductionData!E:E, C{i+1}+1) * '
        f'(SUMPRODUCT(--(B{i+1}="木"), 1 + (0.25*($B$23="是") + IF($B$24>0, $B$24, 0))) + '
        f'SUMPRODUCT(--(B{i+1}="磚"), 1 + (0.25*($B$26="是") + IF($B$27>0, $B$27, 0))) + '
        f'SUMPRODUCT(--(B{i+1}="鐵"), 1 + (0.25*($B$29="是") + IF($B$30>0, $B$30, 0))) + '
        f'SUMPRODUCT(--(B{i+1}="米"), 1 + (0.25*($B$32="是") + (0.25*($B$33="是") + IF($B$34>0, $B$34, 0))))) * '
        f'(SUMPRODUCT(--(B{i+1}="木"), IF($B$25="是", 1.25, 1)) + '
        f'SUMPRODUCT(--(B{i+1}="磚"), IF($B$28="是", 1.25, 1)) + '
        f'SUMPRODUCT(--(B{i+1}="鐵"), IF($B$31="是", 1.25, 1)) + '
        f'SUMPRODUCT(--(B{i+1}="米"), IF($B$35="是", 1.25, 1)))))'
    ))

    # 回本時間計算
    ws.cell(row=i+1, column=8, value=(
        f'=IFERROR(IF(E{i+1} - G{i+1} <> 0, F{i+1} / (E{i+1} - G{i+1}), "無法計算"), "計算錯誤")'
    ))

    # 將所有數據行的對齊方式設置為靠右
    for col in range(1, 9):
        ws.cell(row=i+1, column=col).alignment = Alignment(horizontal="right")

# 合計行
ws.cell(row=20, column=4, value="合計：").alignment = Alignment(horizontal="right")
ws.cell(row=20, column=5, value=f"=SUM(E2:E19)").alignment = Alignment(horizontal="right")
ws.cell(row=20, column=6, value=f"=SUM(F2:F19)").alignment = Alignment(horizontal="right")

# 加乘控制
ws.merge_cells('A22:C22')
ws['A22'] = "加乘控制"

yes_no_list = '"是,否"'
ws['A23'] = "木 建築加乘"
ws['B23'] = "是"
ws['A24'] = "木 綠洲加乘"
ws['B24'] = 0.0
ws['A25'] = "木 金幣加乘"
ws['B25'] = "是"

# 將加乘控制部分的數值列置右
ws['B23'].alignment = Alignment(horizontal="right")
ws['B24'].alignment = Alignment(horizontal="right")
ws['B25'].alignment = Alignment(horizontal="right")

dv_yes_no = DataValidation(type="list", formula1=yes_no_list, showDropDown=True)
ws.add_data_validation(dv_yes_no)
dv_yes_no.add(ws['B23'])
dv_yes_no.add(ws['B25'])

dv_wood_oasis = DataValidation(type="list", formula1='"0.0,0.25,0.50,0.75,1.00,1.25,1.50"', showDropDown=True)
ws.add_data_validation(dv_wood_oasis)
dv_wood_oasis.add(ws['B24'])

ws['A26'] = "磚 建築加乘"
ws['B26'] = "是"
ws['A27'] = "磚 綠洲加乘"
ws['B27'] = 0.0
ws['A28'] = "磚 金幣加乘"
ws['B28'] = "是"

# 將加乘控制部分的數值列置右
ws['B26'].alignment = Alignment(horizontal="right")
ws['B27'].alignment = Alignment(horizontal="right")
ws['B28'].alignment = Alignment(horizontal="right")

dv_yes_no.add(ws['B26'])
dv_yes_no.add(ws['B28'])

dv_brick_oasis = DataValidation(type="list", formula1='"0.0,0.25,0.50,0.75,1.00,1.25,1.50"', showDropDown=True)
ws.add_data_validation(dv_brick_oasis)
dv_brick_oasis.add(ws['B27'])

ws['A29'] = "鐵 建築加乘"
ws['B29'] = "是"
ws['A30'] = "鐵 綠洲加乘"
ws['B30'] = 0.0
ws['A31'] = "鐵 金幣加乘"
ws['B31'] = "是"

# 將加乘控制部分的數值列置右
ws['B29'].alignment = Alignment(horizontal="right")
ws['B30'].alignment = Alignment(horizontal="right")
ws['B31'].alignment = Alignment(horizontal="right")

dv_yes_no.add(ws['B29'])
dv_yes_no.add(ws['B31'])

dv_iron_oasis = DataValidation(type="list", formula1='"0.0,0.25,0.50,0.75,1.00,1.25,1.50"', showDropDown=True)
ws.add_data_validation(dv_iron_oasis)
dv_iron_oasis.add(ws['B30'])

ws['A32'] = "米 建築加乘1"
ws['B32'] = "是"
ws['A33'] = "米 建築加乘2"
ws['B33'] = "是"
ws['A34'] = "米 綠洲加乘"
ws['B34'] = 1.5
ws['A35'] = "米 金幣加乘"
ws['B35'] = "是"

# 將加乘控制部分的數值列置右
ws['B32'].alignment = Alignment(horizontal="right")
ws['B33'].alignment = Alignment(horizontal="right")
ws['B34'].alignment = Alignment(horizontal="right")
ws['B35'].alignment = Alignment(horizontal="right")

dv_yes_no.add(ws['B32'])
dv_yes_no.add(ws['B33'])
dv_yes_no.add(ws['B35'])

dv_rice_oasis = DataValidation(type="list", formula1='"0.0,0.25,0.50,0.75,1.00,1.25,1.50"', showDropDown=True)
ws.add_data_validation(dv_rice_oasis)
dv_rice_oasis.add(ws['B34'])

# 保存 Excel 檔案
wb.save("資源產能試算.xlsx")
